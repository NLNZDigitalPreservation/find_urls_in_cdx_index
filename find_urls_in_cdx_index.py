from openpyxl import load_workbook, Workbook
import requests, datetime, sys, time

# Take the input file path from argument [1] and the output folder path from argument [2]
try:
    input_workbook = load_workbook(filename=sys.argv[1])
except:
    print('Input file not found')
    exit()

if len(sys.argv) < 4:
    print('Please include all arguments <path to input spreadsheet> <path to output directory> <cdx index url>')
    exit()

output_filepath = sys.argv[2] + '/urls_found_in_cdx_index.xlsx'
cdx_index = sys.argv[3]

# Create a new spreadsheet
output_workbook = Workbook()
worksheet_1 = output_workbook.active
worksheet_1.title = 'Found URLs'
worksheet_1.append(['URL', 'Occurrence 1', 'Occurrence 2', 'Occurrence 3'])


def format_date(timestamp):
    return datetime.datetime.strptime(str(timestamp), '%Y%m%d%H%M%S').strftime('%d/%m/%Y %H:%M:%S')


urls = []
total_urls, total_urls_duplicates_removed, found_urls, max_url_length = 0, 0, 0, 0

print('Finding unique URLs')
for sheet in input_workbook.worksheets:
    for cell in sheet['E']:
        url = cell.value.strip()
        total_urls += 1
        if url not in urls:
            urls.append(url)

urls.sort()
total_urls_duplicates_removed = len(urls)

for url in urls:
    # Check each URL against the CDX index
    params = (
        ('url', url),
        ('output', 'json'),
        ('limit', 3),
        ('sort', 'reverse'),
    )
    try:
        response = requests.get(cdx_index, params=params).json()
    except:
        print('Unable to connect to CDX Index')
        exit()

    # If it exists in the index add it and the three latest occurrence dates to the output spreadsheet
    if len(response) > 0:
        print('Adding found URL ' + url + ' to output spreadsheet')
        date1 = format_date(response[0][1])
        date2 = ''
        date3 = ''
        if len(response) > 1:
            date2 = format_date(response[1][1])
        if len(response) > 2:
            date3 = format_date(response[2][1])

        if len(url) > max_url_length:
            max_url_length = len(url)

        found_urls += 1
        worksheet_1.append([url, date1, date2, date3])

        # Slow down the requests so the server isn't bombarded
        time.sleep(.25)


worksheet_1.append(['Total input URLs: ' + str(total_urls)])
worksheet_1.append(['Total input URLs with duplicates removed: ' + str(total_urls_duplicates_removed)])
worksheet_1.append(['Total found URLs: ' + str(found_urls)])

# Resize the output columns
worksheet_1.column_dimensions['A'].width = max_url_length
worksheet_1.column_dimensions['B'].width = 25
worksheet_1.column_dimensions['C'].width = 25
worksheet_1.column_dimensions['D'].width = 25

# Save the output spreadsheet
try:
    output_workbook.save(output_filepath)
except:
    print('Unable to save output spreadsheet, check the filepath')
    exit()

print(
        '',
        '',
        'Total input URLs: ' + str(total_urls),
        'Total unique input URLs: ' + str(total_urls_duplicates_removed),
        'Total found URLs added to output spreadsheet: ' + str(found_urls),
        'Output spreadsheet created at ' + output_filepath,
        sep='\n'
    )
